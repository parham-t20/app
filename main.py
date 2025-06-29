import socket
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.anchorlayout import AnchorLayout
from kivy.uix.screenmanager import ScreenManager, Screen
from pytz import timezone
from datetime import datetime
from functools import partial
from kivy.clock import Clock
from openpyxl import load_workbook, Workbook
import time
import os

country_timezones = {
    "iran": "Asia/Tehran",
    "usa": "America/New_York",
    "germany": "Europe/Berlin",
    "china": "Asia/Shanghai",
    "india": "Asia/Kolkata",
    "japan": "Asia/Tokyo",
    "russia": "Europe/Moscow",
    "uae": "Asia/Dubai",
    "brazil": "America/Sao_Paulo",
    "uk": "Europe/London"
}

EXCEL_PATH = "/sdcard/Documents/password.xlsx"

class UserPass(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
            self.password = ws["A1"].value
        else:
            self.password = "0000"

        box = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=22, padding=[10, 30, 10, 10])
        self.lbl = Label(text='Type Password', font_size=30, size_hint_y=None, height=60)
        self.input_password = TextInput(multiline=False, size_hint_y=None, height=50, font_size=30)
        btn_ok = Button(text="Ok", size_hint_y=None, height=80)

        box.add_widget(self.lbl)
        box.add_widget(self.input_password)
        box.add_widget(btn_ok)
        box.bind(minimum_height=box.setter('height'))

        btn_ok.bind(on_press=partial(self.change_screen, 'main'))

        anchor = AnchorLayout(anchor_y='top')
        anchor.add_widget(box)
        self.add_widget(anchor)

    def change_screen(self, screen_name, instance):
        if self.input_password.text == self.password:
            self.lbl.color = (1, 1, 1, 1)
            self.manager.current = screen_name
            self.input_password.text = ""
        else:
            self.lbl.text = "Not Found Password"
            self.lbl.color = (1, 0, 0, 1)

class MainScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        box = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=20, padding=10)
        box.bind(minimum_height=box.setter('height'))

        btn_connect = Button(text='Connect To Device', size_hint_y=None, height=80)
        btn_time_zone = Button(text="Time Zone Finder", size_hint_y=None, height=80)
        btn_op = Button(text="Option", size_hint_y=None, height=80)

        btn_connect.bind(on_press=partial(self.change_screen, 'device'))
        btn_time_zone.bind(on_press=partial(self.change_screen, 'timezone'))
        btn_op.bind(on_press=partial(self.change_screen, 'notebook'))

        box.add_widget(btn_connect)
        box.add_widget(btn_time_zone)
        box.add_widget(btn_op)

        anchor = AnchorLayout(anchor_y='top')
        anchor.add_widget(box)
        self.add_widget(anchor)

    def change_screen(self, screen_name, instance):
        self.manager.current = screen_name

class DeviceScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        box = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=22, padding=[10, 30, 10, 10])
        box.bind(minimum_height=box.setter('height'))

        self.lbl = Label(text='IP Address', font_size=30, size_hint_y=None, height=60)
        self.txt_input_ip = TextInput(text='192.168.', multiline=False, size_hint_y=None, height=50, font_size=30)
        btn_connect = Button(text='Connect Device', size_hint_y=None, height=80)
        btn_back = Button(text='Back', size_hint_y=None, height=80)

        btn_connect.bind(on_press=self.go_to_send)
        btn_back.bind(on_press=partial(self.change_screen, 'main'))

        box.add_widget(self.lbl)
        box.add_widget(self.txt_input_ip)
        box.add_widget(btn_connect)
        box.add_widget(btn_back)

        anchor = AnchorLayout(anchor_y='top')
        anchor.add_widget(box)
        self.add_widget(anchor)

    def change_screen(self, screen_name, instance):
        self.manager.current = screen_name

    def go_to_send(self, instance):
        ip_text = self.txt_input_ip.text
        if ip_text.count('.') != 3 or len(ip_text) < 10:
            self.lbl.text = "Error"
            self.lbl.color = (1, 0, 0, 1)
        else:
            self.lbl.text = "IP Address"
            self.lbl.color = (1, 1, 1, 1)
            send_screen = self.manager.get_screen('send')
            send_screen.ip = ip_text
            self.manager.current = 'send'

class SendToDeviceScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.ip = ""

        box = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=22, padding=[10, 30, 10, 10])
        box.bind(minimum_height=box.setter('height'))

        lbl = Label(text='Your Message', font_size=30, size_hint_y=None, height=60)
        self.txt_input_Message = TextInput(multiline=False, size_hint_y=None, height=50, font_size=30)
        btn_send = Button(text='Send Message', size_hint_y=None, height=80)
        btn_back = Button(text='Back', size_hint_y=None, height=80)

        btn_send.bind(on_press=self.send_message)
        btn_back.bind(on_press=partial(self.change_screen, 'device'))

        box.add_widget(lbl)
        box.add_widget(self.txt_input_Message)
        box.add_widget(btn_send)
        box.add_widget(btn_back)

        anchor = AnchorLayout(anchor_y='top')
        anchor.add_widget(box)
        self.add_widget(anchor)

    def change_screen(self, screen_name, instance):
        self.manager.current = screen_name

    def send_message(self, instance):
        msg = self.txt_input_Message.text
        try:
            client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            client.connect((self.ip, 12345))
            client.send(msg.encode('utf-8'))
            response = client.recv(1024).decode('utf-8')
            print("Response from server:", response)
            client.close()
        except Exception as e:
            print("Connection error:", e)

class TimeZoneScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        box = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=22, padding=[10, 30, 10, 10])
        box.bind(minimum_height=box.setter('height'))

        self.time_lbl = Label(text=f'Country name Time is 00:00:00', font_size=30, size_hint_y=None, height=60)
        self.txt_input_time = TextInput(multiline=False, size_hint_y=None, height=50, font_size=30)
        btn_find = Button(text='Find', size_hint_y=None, height=80)
        btn_back = Button(text='Back', size_hint_y=None, height=80)

        box.add_widget(self.time_lbl)
        box.add_widget(self.txt_input_time)
        box.add_widget(btn_find)
        box.add_widget(btn_back)

        btn_find.bind(on_press=self.finder)
        btn_back.bind(on_press=partial(self.change_screen, 'main'))

        anchor = AnchorLayout(anchor_y='top')
        anchor.add_widget(box)
        self.add_widget(anchor)

    def change_screen(self, screen_name, instance):
        self.manager.current = screen_name

    def finder(self, instance):
        try:
            country = self.txt_input_time.text.strip().lower()
            if country in country_timezones:
                tz_name = country_timezones[country]
                tz = timezone(tz_name)
                now = datetime.now(tz)
                self.time_lbl.color = (1, 1, 1, 1)
                self.time_lbl.text = f"Country {country.title()} Time is {now.strftime('%H:%M:%S')}"
            else:
                self.time_lbl.text = "Country not found"
                self.time_lbl.color = (1, 0, 0, 1)
        except Exception as e:
            self.time_lbl.text = "Error occurred"
            self.time_lbl.color = (1, 0, 0, 1)
            print("Error in finder:", e)

class NoteBook(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        box5 = BoxLayout(orientation='vertical', size_hint=(1, None), spacing=22, padding=[10, 30, 10, 10])
        box5.bind(minimum_height=box5.setter('height'))

        self.input_password_save = TextInput(hint_text="Password", multiline=False, size_hint_y=None, height=50, font_size=30)
        self.input_password_save1 = TextInput(hint_text="Password Confirm", multiline=False, size_hint_y=None, height=50, font_size=30)
        self.btn_save = Button(text='Save Password', size_hint_y=None, height=80)

        self.btn_save.bind(on_press=self.save)

        box5.add_widget(self.input_password_save)
        box5.add_widget(self.input_password_save1)
        box5.add_widget(self.btn_save)

        anchor = AnchorLayout(anchor_y='top')
        anchor.add_widget(box5)
        self.add_widget(anchor)

    def save(self, instance):
        if self.input_password_save.text == self.input_password_save1.text:
            self.save_to_excel()
            self.btn_save.text = "Saved"
            self.btn_save.color = (0, 1, 0, 1)
            Clock.schedule_once(self.reset_btn, 1.5)
        else:
            self.btn_save.text = "Error"
            self.btn_save.color = (1, 0, 0, 1)
            Clock.schedule_once(self.reset_btn, 1.5)

    def reset_btn(self, dt):
        self.btn_save.text = "Save Password"
        self.btn_save.color = (1, 1, 1, 1)

    def save_to_excel(self):
        folder = "/sdcard/Documents"
        if not os.path.exists(folder):
            os.makedirs(folder)
        password = self.input_password_save.text

        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        ws["A1"] = password
        wb.save(EXCEL_PATH)

class MyApp(App):
    def build(self):
        sm = ScreenManager()
        sm.add_widget(UserPass(name='userpass'))
        sm.add_widget(MainScreen(name='main'))
        sm.add_widget(DeviceScreen(name='device'))
        sm.add_widget(SendToDeviceScreen(name='send'))
        sm.add_widget(TimeZoneScreen(name='timezone'))
        sm.add_widget(NoteBook(name='notebook'))
        return sm

if __name__ == '__main__':
    MyApp().run()
